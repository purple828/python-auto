'''
读取excel文件内容并写入数据库

'''
from openpyxl import load_workbook
import pymysql
import configparser
import os

#项目路径
print('---------',os.path.realpath(__file__))
print('*******',os.path.split(os.path.realpath(__file__)))
root_dir = os.path.split(os.path.realpath(__file__))[0]
config_file_path = os.path.join(root_dir,'config.ini')

def get_config_info(section,options):
    infos = []
    # 根据传入的section获取对应的value section为ini配置文件中用[]标识的内容
    config = configparser.ConfigParser()
    config.read(config_file_path)
    if isinstance(options, list):
        if options:
            for option in options:
                try:
                    value = config.get(section=section, option=option)
                    print('value---', value)
                    infos.append(value)
                except Exception as e:
                    print('获取{}的值失败：{}'.format(option,e))

    else:
        try:
            value = config.get(section=section, option=options)
            print('value---', value)
            infos.append(value)
        except Exception as e:
            print('获取{}的值失败：{}'.format(options, e))

    return infos

res = get_config_info('mysql',['host','user','password'])

# conn = pymysql.connect("localhost","root","123456aa",use_unicode=True, charset="utf8")
conn = pymysql.connect(res[0],res[1],res[2],use_unicode=True, charset="utf8")

conn.autocommit(1)

cur = conn.cursor()

#创建数据库exceltest
name="exceltest"
db_sql = "create database if not exists %s"%name
try:
    cur.execute(db_sql)
except:
    print('数据库{}创建失败'.format(name))

# 选择要操作的数据库
conn.select_db(name)
table_name = "douban_movie"
table_sql = '''
            create table if not exists %s(
              id mediumint not null auto_increment,
              name varchar(50),
              score decimal,
              comment varchar(50),
              time varchar(50),
              url varchar(100),
              PRIMARY KEY (`id`)
            )
'''%table_name

cur.execute(table_sql)

#读取excel文件内容
#获取excel工作页面（地址可以是相对路径或绝对路径）
wb2 = load_workbook('豆瓣电影.xlsx')
#获取所有sheet的名称
sheetnames = wb2.get_sheet_names()
print('sheetnames----------',sheetnames)
#得到的是一个Worksheet对象
ws = wb2.get_sheet_by_name(sheetnames[0])
print('ws-------------',ws)
print('{} title:{}'.format(ws,ws.title))
print('{} rows:{}'.format(ws,ws.max_row))
print('{} cols:{}'.format(ws,ws.max_column))

del_sql = "delete from {}".format(table_name)
cur.execute(del_sql)

# for i in range(2,ws.max_row+1):
#     value1s = []
#     for j in range(1,ws.max_column+1):
#         v = ws.cell(row=i, column=j).value
#         value1s.append(v)
#     #将excel表的每一行作为一条记录插入数据库
#     print('values------',value1s)
#     # value1s.insert(0,table_name)
#     # print('values------', value1s)
#     insert_sql = "insert into douban_movie(name,score,comment,time,url) values(%s,%s,%s,%s,%s)"
#     print('====================',tuple(value1s))
#     cur.execute(insert_sql,tuple(value1s))




#获取excel文件中的所有sheet对象
count = 1
for sheet in wb2:
    print('sheet-------------------',sheet)
    #每一行为一个cell对象
    num = 1
    for cell in sheet:
        num += 1
        # print('cell{}----------------------------'.format(cell))
        # max_col = sheet.max_column
        # for i in range(0,max_col+1):
        value1 = (cell[0].value,cell[1].value,cell[2].value,cell[3].value,cell[4].value)
        print('*************',value1)
        if num > 2:
            insert_sql = "insert into douban_movie(name,score,comment,time,url) values(%s,%s,%s,%s,%s)"
            cur.execute(insert_sql,value1)












